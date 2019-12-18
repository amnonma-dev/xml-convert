import xml.etree.ElementTree as ET
import csv 
import os
import random
import openpyxl as exl
import easygui
import time


def replace_data(field,value):
    if value is not None:
        field.text= str(value)

def RepresentsInt(s):
    if s is None:
        return False
    try: 
        int(s)
        return True
    except ValueError:
        return False

def errhandler(option_arg):    
    print("invalid option: "+str(option_arg))

def convert_xml(is_shipment):
    
    fields_map={'AddressLine1':4,'AddressLine3City':6,'AddressLine4State':7,'CompanyName':3,'CountryCode':9,'EndContactEmail':10,'EndContactPhone':11,\
                'FirstName':1,'OrderNumber':0,'ProductUnitWeight':14,'Postcode':8,'ProductHarmonizedCode':12,'ProductDescription':13,'ProductUnitValue':15,\
                'ProductQuantity':16,'ProductItemOrigin':17,'Length':21,'Width':22,'Height':23,'ParcelWeight':19}
    # 'LastName':2,
    folder_name='outfiles/Deliveries/'
    xml_name='4log.xml'
    max_column=24

    if is_shipment:
        fields_map['Fourlogref']=24
        fields_map['AddrValid']=25
        folder_name='outfiles/Shipments/'
        xml_name='4log_createshipment.xml'
        max_column=26

    with open('country-codes.txt') as csvfile:
        reader = csv.reader(csvfile,delimiter='\t')
        country_map = {rows[0]:rows[1] for rows in reader}


    outdir = os.path.dirname(folder_name)
    os.makedirs(outdir,exist_ok=True)

    folder = os.path.dirname(folder_name)
    for the_file in os.listdir(folder):
        file_path = os.path.join(folder, the_file)
        try:
            if os.path.isfile(file_path):
                os.unlink(file_path)
            #elif os.path.isdir(file_path): shutil.rmtree(file_path)
        except Exception as e:
            print(e)

    ns={'SOAP-ENV':'http://schemas.xmlsoap.org/soap/envelope/','ns1':'http://tempuri.org/','ns2':'http://schemas.datacontract.org/2004/07/ShowbalAPI'}
    ET.register_namespace('SOAP-ENV','http://schemas.xmlsoap.org/soap/envelope/')

    tree = ET.parse(xml_name)
    root=tree.getroot()

    orders_file_name=easygui.fileopenbox(msg="select orders file to upload",)
    if orders_file_name is None:
        exit()
    try:
        wb = exl.load_workbook(filename = orders_file_name,data_only=True)
        active_sheet=wb[wb.sheetnames[0]]
        # print(active_sheet.max_row)
    except:
        raise Exception("Problem reading orders file")
    #handle username password
    vendors = {
        1: "JWG",
        2: "CAROLINA"
    }
    while True:
        for key in sorted(vendors):
            print (key, '=>', vendors[key])
        selected_vendor = input("please select an option from the list:    ")
        try:
            selected_vendor= int(selected_vendor)
        except:
            pass

        if selected_vendor==1:
            break
        elif selected_vendor==2:
            replace_data(root.find('.//ns1:userName',ns),'CAROLINAusername')
            replace_data(root.find('.//ns1:password',ns),'CAROLINApassword')
            break
        else:
            errhandler(selected_vendor)
            continue

    line_number=2
    for row in active_sheet.iter_rows(2,active_sheet.max_row,1,max_column):
        if row[0].value is None:
            continue
        for field,field_map in fields_map.items():
            if field=='EndContactPhone':
                phone=row[field_map].value
                if phone is not None:
                    phone="\'"+str(phone)
                replace_data(root.find('.//ns2:'+field,ns),phone)   

            elif field=='CountryCode':
                country_name=str(row[field_map].value)
                if country_name in country_map.keys():
                    country_code= country_map[country_name]
                elif country_name in country_map.values():
                    country_code= country_name
                else:
                    print("Warning- no match for country:"+country_name+"(line "+str(line_number)+")")
                    country_code= None

                replace_data(root.find('.//ns2:'+field,ns),country_code)   

            elif field=='ProductUnitValue':
                unit_value=str(row[field_map].value)
                ind=0
                for char in unit_value:
                    if  RepresentsInt(char):
                        break
                    ind+=1
                unit_value=unit_value[ind:]
                replace_data(root.find('.//ns2:'+field,ns),unit_value)   

            elif field=='ProductUnitWeight':
                weight=row[field_map].value
                if country_code!='US':
                    try:
                        weight= float(weight)*1000
                    except:
                        pass
                replace_data(root.find('.//ns2:'+field,ns),weight) 
            elif field=='FirstName':
                full_name= str(row[field_map].value).strip()
                # while '  ' in full_name:
                #     full_name = full_name.replace('  ', ' ')
                if ' ' in full_name:
                    first_name= full_name.split(' ')[0]
                    last_name= ' '.join(full_name.split(' ')[1:])
                    replace_data(root.find('.//ns2:LastName',ns),last_name) 
                else:
                    first_name=full_name
                replace_data(root.find('.//ns2:'+field,ns),first_name) 
                
            else:
                replace_data(root.find('.//ns2:'+field,ns),row[field_map].value)  
        if country_code=='US':
            root.find('.//ns2:ShipType',ns).text='1'
        else:
            root.find('.//ns2:ShipType',ns).text='2'
        tree.write(file_or_filename=os.path.join(outdir,str(row[0].value)+'.xml'),xml_declaration=True,encoding='UTF-8')
        line_number+=1
    print("processed "+str(line_number-2)+" lines for vendor:"+vendors[selected_vendor])


def main():
    actions = {
            1: "convert delivery",
            2: "convert shipment",
            3: "quit"
        }
    
    while True:
        for key in sorted(actions):
            print (key, '=>', actions[key])
        selectedaction = input("please select an option from the list:    ")
        try:
            selectedaction= int(selectedaction)
        except:
            pass
            # errhandler(selectedaction)
            # continue

        if selectedaction==1:
            convert_xml(False)
            time.sleep(3)
        elif selectedaction==2:
            convert_xml(True)
            time.sleep(3)
        elif selectedaction==3:
            exit()
        else:
            errhandler(selectedaction)
            continue

main()